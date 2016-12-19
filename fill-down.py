from openpyxl import load_workbook
import glob


def main():

    for file in list(glob.glob('*.xlsx')):
        print (file)
        wb = load_workbook(file)
        ws = wb['OR']

        for col in ws.iter_cols(min_col=2, max_col=5):
            current_value = None
            for cell in col:
                if cell.value is None:
                    cell.value = current_value
                elif not unicode(cell.value).replace('.', '').isdigit():
                    current_value = cell.value

        new_file = file.replace('.xlsx', 'filled.xlsx')
        wb.save(new_file)

if __name__ == '__main__':
    main()
